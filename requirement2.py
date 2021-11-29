def execute():
    print("Requirement 2 Code")
    import cv2
    import numpy as np
    from matplotlib import pyplot as pt 
    import pytesseract as tess
    from openpyxl import Workbook, load_workbook
    
    tess.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    
    for n in range(1,21): 
        if n < 10: 
            img_num = str(0) + str(n) 
        else: 
            img_num = str(n) 
         
        #read image 
        img_number = img_num + ".png" 
        img_extra = cv2.imread(Dataset/img_number, cv2.IMREAD_GRAYSCALE) 
         
        #convert image to grayscale 
        thresh = 128 
        img = cv2.threshold(img_extra, thresh, 255, cv2.THRESH_BINARY)[1] 
         
        #get the height and width of image 
        hImg,wImg = img.shape 
                
        #make the text clearer
        sE = np.array([ [0,1,0],[1,1,1],[0,1,0] ])
        sE = sE.astype(np.uint8)
        rSE = np.rot90(sE,2)
        img = cv2.erode(img,rSE,iterations=3)
        img = cv2.dilate(img,rSE,iterations=2)
        
        def calcLength(startPt,endPt) :
            legX = endPt[0] - startPt[0]
            legY = endPt[1] - startPt[1]
            hypotenuse = np.sqrt(np.power(legX,2) + np.power(legY,2))
            return hypotenuse
        
        def calcMidpoint(startPt,endPt) :
            midPt = [0,0]
            midPt[0] = (startPt[0] + endPt[0])/2
            midPt[1] = (startPt[1] + endPt[1])/2
            return midPt
    
        dict_output = tess.image_to_data(img, output_type=tess.Output.DICT)
        length = len(dict_output['level'])
        
        for m in range (length):
            if (dict_output['text'][m] == 'STATUS:'):
        
                coordinate = []
                for a in range (length):
                    x_a = dict_output['left'][a]
                    y_a = dict_output['top'][a]
                    w_a = dict_output['width'][a]
                    h_a = dict_output['height'][a]
                    t_a = dict_output['text'][a]
                    if (t_a != ""):
                        list1 = [x_a,y_a,w_a,h_a,t_a]
                        coordinate.append(list1)
                # find the coordinates of every other text
                point_list = []
                for sublist in coordinate:
                    if (sublist[-1] == 'STATUS:'):
                        start1 = [sublist[0],sublist[1]]
                        start2 = [sublist[0]+sublist[2],sublist[1]+sublist[3]]
                        midStart = calcMidpoint(start1,start2)
                    else:
                        point = [sublist[0],sublist[1]]
                        point_list.append(point)
    
                # find the minimum distance
                distance_list = []
                for i in point_list:
                    # print(i)
                    distance = calcLength(midStart,i)
                    distance_list.append(distance)
                      
                # get minimum distance
                min_distance = min(distance_list)
                index = distance_list.index(min_distance)
                for sublist in coordinate:
                    if (sublist[0] == point_list[index][0]) and (sublist[1] == point_list[index][1]):
                        content = sublist[-1]
                        print(content)
                    
   #save drawing number into excel sheet 
    if (img_num == '01'): 
        wb = load_workbook('DIP assignment.xlsx')
        wb.create_sheet('Status')
    else : 
        wb = load_workbook('DIP assignment.xlsx')
        ws = wb['Status']
        ws.append([str(n),'Status:',content])
        wb.save('DIP assignment.xlsx')

