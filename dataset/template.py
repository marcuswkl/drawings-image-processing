# -*- coding: utf-8 -*-
"""
Created on Mon Nov 29 07:35:53 2021

@author: Asus
"""

# -*- coding: utf-8 -*-
import pytesseract as tess
import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook

tess.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

#loop through all 20 image
for n in range(1,21):
    if n < 10:
        img_num = str(0) + str(n)
    else:
        img_num = str(n)
    
    #read image
    img_number = img_num + ".png"
    img_extra = cv2.imread(img_number, cv2.IMREAD_GRAYSCALE)
    
    #convert image to grayscale
    thresh = 128
    img = cv2.threshold(img_extra, thresh, 255, cv2.THRESH_BINARY)[1]
    
    #get the height and width of image
    hImg,wImg = img.shape
    
    #function to calculate midpoint
    def calcMidpoint(startPt,endPt) :
        midPt = [0,0]
        midPt[0] = (startPt[0] + endPt[0])/2
        midPt[1] = (startPt[1] + endPt[1])/2
        return midPt
    
    #function to calculate length between two points
    def calcLength(startPt,endPt) :
        legX = endPt[0] - startPt[0]
        legY = endPt[1] - startPt[1]
        hypotenuse = np.sqrt(np.power(legX,2) + np.power(legY,2))
        return hypotenuse
    
    #turn the data into text, loop through all data and find the distance between 'NUMBER' text and 
    #also each data in the list and save them into a list
    def dataToText(img) :
        boxes = tess.image_to_data(img) 
        distance = 0
        distance_list = []
        item_list = []
        for x,b in enumerate(boxes.splitlines()):
            if x != 0:
                b = b.split()
                if len(b) == 12:
                    x,y,w,h = int(b[6]),int(b[7]),int(b[8]),int(b[9])
                    cv2.rectangle(img,(x,y),(w+x,h+y),(0,0,255),3)
                    cv2.putText(img,b[11],(x,y),cv2.FONT_HERSHEY_COMPLEX,1,(50,50,255),2)
                    if ((b[11] == 'NUMBER:') or (b[11] == 'NO.:') or (b[11] == 'NO.::') or (b[11] == 'NO:')):
                        startPt = [x,y]
                        endPt = [w+x,y+h]
                        for x,b in enumerate(boxes.splitlines()):
                                if x != 0:
                                    b = b.split()
                                    if ((len(b) == 12) and ((b[11] != 'NUMBER:') or (b[11] != 'NO.:') or (b[11] != 'NO.::') or (b[11] == 'NO:')) and ((b[11] != 'DRAWING:') or (b[11] != 'DRAWING') or (b[11] != '|DRAWING'))):
                                        x_a,y_a,w_a,h_a = int(b[6]),int(b[7]),int(b[8]),int(b[9])
                                        startPt_a = [x_a, y_a]
                                        endPt_a = [x_a + w_a, y_a + h_a]
                                        mp = calcMidpoint(startPt, endPt)
                                        mp_a = calcMidpoint(startPt_a, endPt_a)
                                        distance = calcLength(mp,mp_a)
                                        if (distance > 0):
                                            item_list.append(b[11])
                                            distance_list.append(distance)
        return distance_list, item_list
                
    #make the text more readable
    def morphText(img,it1,it2) :
        sE = np.array([ [0,1,0],[1,1,1],[0,1,0] ])
        sE = sE.astype(np.uint8)
        rSE = np.rot90(sE,2)
        img = cv2.erode(img,rSE,iterations=it1)
        img = cv2.dilate(img,rSE,iterations=it2)
        return img
    
    #remove horizontal lines
    #idea obtained from website
    def removeHori(img3):
        horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40,1))
        remove_horizontal = cv2.morphologyEx(threshImg, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
        contoursImg = cv2.findContours(remove_horizontal, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contoursImg = contoursImg[0] if len(contoursImg) == 2 else contoursImg[1]
        for c in contoursImg:
            cv2.drawContours(img3, [c], -1, (255,255,255), 5)
        return img3

    #remove long vertical lines
    #idea obtained from website
    def removeLongVerti(img3):
        vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1,40))
        remove_vertical = cv2.morphologyEx(threshImg, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
        contoursImg = cv2.findContours(remove_vertical, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contoursImg = contoursImg[0] if len(contoursImg) == 2 else contoursImg[1]
        for c in contoursImg:
            cv2.drawContours(img3, [c], -1, (255,255,255), 5)
        return img3

    #remove short vertical lines
    #idea obtained from website
    def removeShortVerti(img3):
        vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1,20))
        remove_vertical = cv2.morphologyEx(threshImg, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
        contoursImg = cv2.findContours(remove_vertical, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contoursImg = contoursImg[0] if len(contoursImg) == 2 else contoursImg[1]
        for c in contoursImg:
            cv2.drawContours(img3, [c], -1, (255,255,255), 5)
        return img3
    
    #turn the data into text, add all text into a list
    def dataToText2(img):
        boxes = tess.image_to_data(img) 
        data_list = []
        for x,b in enumerate(boxes.splitlines()):
            if x != 0:
                b = b.split()
                if len(b) == 12:
                    x,y,w,h = int(b[6]),int(b[7]),int(b[8]),int(b[9])
                    cv2.rectangle(img,(x,y),(w+x,h+y),(0,0,255),3)
                    cv2.putText(img,b[11],(x,y),cv2.FONT_HERSHEY_COMPLEX,1,(50,50,255),2)
                    if ((b[11] == 'NUMBER:') or (b[11] == 'NO.:') or (b[11] == 'NO.::') or (b[11] == 'NO:')):
                        for x,b in enumerate(boxes.splitlines()):
                                if x != 0:
                                    b = b.split()
                                    if ((len(b) == 12)):
                                        if((len(data_list)==0)):
                                            data_list.append(b[11])
                                        elif((b[11] != data_list[0]) and (len(data_list)>= 1)):
                                            data_list.append(b[11])
                                        else:
                                            break
        return data_list
        
    #first condition: data are not bounded by text
    img = morphText(img,3,2)
    distance_list, item_list = dataToText(img)
    
    #find them item with the smallest distance to "NUMBER"
    if len(distance_list) != 0:
        smallest_distance = min(distance_list)
        index = distance_list.index(smallest_distance)
        drawing_number = item_list[index]
        if((drawing_number.isalpha()) or (drawing_number.isnumeric()) or (not len(drawing_number) > 15)):
            item_list.remove(item_list[index])
            distance_list.remove(distance_list[index])
            smallest_distance = min(distance_list)
            index = distance_list.index(smallest_distance)
            drawing_number = item_list[index]
        if((drawing_number.isalpha()) or (drawing_number.isnumeric()) or (not len(drawing_number) > 15)):
            item_list.remove(item_list[index])
            distance_list.remove(distance_list[index])
            smallest_distance = min(distance_list)
            index = distance_list.index(smallest_distance)
            drawing_number = item_list[index]
    
    #second condition: for image 06.png
    #different morphological operation iteration only
    if((drawing_number.isalpha()) or (drawing_number.isnumeric()) or (len(drawing_number) < 15)):
        
        img2 = img_extra.copy()
        #make the text clearer
        img2 = morphText(img2,1,2)
        
        boxes = tess.image_to_data(img2) 
        distance = 0
        distance_list = []
        item_list = []
        for x,b in enumerate(boxes.splitlines()):
            if x != 0:
                b = b.split()
                if len(b) == 12:
                    x,y,w,h = int(b[6]),int(b[7]),int(b[8]),int(b[9])
                    cv2.rectangle(img,(x,y),(w+x,h+y),(0,0,255),3)
                    cv2.putText(img,b[11],(x,y),cv2.FONT_HERSHEY_COMPLEX,1,(50,50,255),2)
                    if ((b[11] == 'NUMBER:') or (b[11] == 'NO.:') or (b[11] == 'NO.::')):
                        startPt = [x,y]
                        endPt = [w+x,y+h]
                        for x,b in enumerate(boxes.splitlines()):
                                if x != 0:
                                    b = b.split()
                                    if ((len(b) == 12) and ((b[11] != 'NUMBER:') or (b[11] != 'NO.:') or (b[11] != 'NO.::')) and ((b[11] != 'DRAWING:') or (b[11] != 'DRAWING'))):
                                        midPt = [0,0]
                                        x_a,y_a,w_a,h_a = int(b[6]),int(b[7]),int(b[8]),int(b[9])
                                        startPt_a = [x_a, y_a]
                                        endPt_a = [x_a + w_a, y_a + h_a]
                                        mp = calcMidpoint(startPt, endPt)
                                        mp_a = calcMidpoint(startPt_a, endPt_a)
                                        distance = calcLength(mp,mp_a)
                                        if (distance > 0):
                                            item_list.append(b[11])
                                            distance_list.append(distance)
                        
                        smallest_distance = min(distance_list)
                        index = distance_list.index(smallest_distance)
                        drawing_number = item_list[index]
                        if((drawing_number.isalpha()) or (drawing_number.isnumeric()) or (not len(drawing_number) > 10)):
                            item_list.remove(item_list[index])
                            distance_list.remove(distance_list[index])
                            smallest_distance = min(distance_list)
                            index = distance_list.index(smallest_distance)
                            drawing_number = item_list[index]
    
    #third condition: text is surrounded by boxes
    if((drawing_number.isalpha()) or (drawing_number.isnumeric()) or (len(drawing_number) < 15)):
        
        #use thresholding to get the binary version of the image
        #idea obtained from website
        img_extra = cv2.imread(img_number)
        img3 = img_extra.copy()
        gray = cv2.cvtColor(img_extra, cv2.COLOR_BGR2GRAY)
        threshImg = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        
        img3 = removeHori(img3)
        
        img3 = removeLongVerti(img3)
        
        img_extra = img3
        img3 = img_extra.copy()
        gray = cv2.cvtColor(img_extra, cv2.COLOR_BGR2GRAY)
        threshImg = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        
        img3 = removeShortVerti(img3)
        
        #get the height and width of img
        hImg,wImg,nlayer = img3.shape
            
        img3 = morphText(img3,3,2)
        
        data_list = dataToText2(img3)
        
        #find the drawing number data by starting after getting 'DRAWING NUMBER'
        #and end when the data list is more than a certain number and ends with an alphabet
        for x,b in enumerate(data_list):
            if((b == 'DRAWING') or (b == 'DRAWING:')):
                if((data_list[x+1] == 'NUMBER:') or (data_list[x+1] == 'NO.:') or (data_list[x+1] == 'NO.::') or (data_list[x+1] == 'NO:')):
                    get_range = x+2
                    for p in range(get_range,len(data_list)):
                        if(data_list[p].isalpha() and ((data_list[p] == 'SU') or (data_list[p] == 'D'))):
                            get_number = p
                            break
        
        drawing_number_list = []                   
        for i in range(get_number,len(data_list)):
            if(i<(8+get_number)):
                drawing_number_list.append(data_list[i])
            elif(data_list[i].isalpha() and (i>(8+get_number))):
                break
            
        drawing_number = '.'.join([str(elem) for elem in drawing_number_list])
    
    #fourth condition: only image 17.png because text is read horizontally and messed up
    if((drawing_number.isalpha()) or (drawing_number.isnumeric()) or (len(drawing_number) < 15) or (len(drawing_number) > 25)):
    
        #convert image4 to grayscale
        image4 = cv2.imread(img_number)
        img4 = image4.copy()
        gray = cv2.cvtColor(image4, cv2.COLOR_BGR2GRAY)
        threshImg = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        
        img4 = removeHori(img4)
        img4 = removeLongVerti(img4)
        
        image4 = img4
        img4 = image4.copy()
        gray = cv2.cvtColor(image4, cv2.COLOR_BGR2GRAY)
        threshImg = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        
        img4 = removeShortVerti(img4)
        
        #crop the drawing to get the specific area only
        img4 = img4[2172:2396,1582:2371]
        
        #get the height and width of image4
        hImg,wImg,nlayer = img4.shape
        
        data_list = dataToText2(img4)
        
        for x,b in enumerate(data_list):
            if((b == 'DRAWING') or (b == 'DRAWING:')):
                if((data_list[x+1] == 'NUMBER:') or (data_list[x+1] == 'NO.:') or (data_list[x+1] == 'NO.::') or (data_list[x+1] == 'NO:')):
                    get_range = x+2
                    for p in range(get_range,len(data_list)):
                        if(data_list[p].isalpha() and ((data_list[p] == 'SU') or (data_list[p] == 'D'))):
                            get_number = p
                            break
        
        drawing_number_list = []
        for i in range(get_number,len(data_list)):
            drawing_number_list.append(data_list[i])
            
        drawing_number = ""
        for i in drawing_number_list:
            drawing_number += str(i) + "."
        drawing_number = drawing_number[:-1]
            
    #save drawing number into excel sheet
    if (img_num == '01'):
        wb = Workbook()
    else :
        wb = load_workbook('DIP assignment.xlsx')
    ws = wb.active
    ws.title = "Drawing Number"
    current_data = 'A' + str(n)
    current_drawing = 'B' + str(n)
    drawing_num = "drawing_" + str(n) + ".png"
    ws[current_drawing] = drawing_number
    ws[current_data] = str(n)
    wb.save('DIP assignment.xlsx')